import json
from pathlib import Path
from openpyxl import load_workbook
import sys
import io
import os
from dotenv import load_dotenv

# Add project root to Python path
project_root = os.path.join(os.path.dirname(__file__), '..', '..')
sys.path.append(project_root)

from src.utils.gcs_utils import get_gcs_handler
from src.utils.data_collector import get_collector

load_dotenv()
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

class FormulaApplier:
    """Apply formula templates to output/main_carriageway.xlsx by replacing {row} placeholders."""
    
    def __init__(self, template_path=None, input_excel_path=None, output_excel_path=None):
        if template_path is None:
            # Look for formula_template.json in project root
            current_dir = Path(__file__).parent
            template_path = current_dir.parent.parent / "formula_template.json"
        
        self.template_path = Path(template_path)
        self.template = self._load_template()
        
        # Initialize session and GCS
        self.session_id = os.getenv('SESSION_ID', 'default')
        self.gcs = get_gcs_handler()
        
        # Determine output filename based on is_merged
        is_merged = os.getenv('IS_MERGED', 'True').lower() == 'true'
        if is_merged:
            self.output_filename = f"{self.session_id}_main_carriageway_and_boq.xlsx"
        else:
            self.output_filename = f"{self.session_id}_main_carriageway.xlsx"
        
        # Handle input/output paths - prefer SESSION_OUTPUT_FILE if available
        if input_excel_path is None:
            # Check for SESSION_OUTPUT_FILE first (local file)
            session_output_file = os.getenv('SESSION_OUTPUT_FILE', '')
            if session_output_file and os.path.exists(session_output_file):
                self.input_excel_path = Path(session_output_file)
                print(f"Using local output file: {session_output_file}")
            else:
                # Fallback: download from GCS (for backward compatibility)
                self.output_gcs_path = self.gcs.get_gcs_path(
                    self.session_id, 
                    self.output_filename, 
                    'output'
                )
                self.input_excel_path = Path(self.gcs.download_to_temp(self.output_gcs_path, suffix='.xlsx'))
                print(f"[GCS] Downloaded output file from GCS: {self.input_excel_path}")
        else:
            self.input_excel_path = Path(input_excel_path)

        if output_excel_path is None:
            # Use same file as input (work on the temp file)
            self.output_excel_path = self.input_excel_path
        else:
            self.output_excel_path = Path(output_excel_path)
        
        # Ensure output directory exists
        self.output_excel_path.parent.mkdir(parents=True, exist_ok=True)
    
    def _load_template(self):
        """Load the formula template from JSON file."""
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        with open(self.template_path, 'r') as f:
            return json.load(f)
    
    def apply_formulas_to_all_data_rows(self, reference_column='D', start_row=7):
        """
        Automatically detect data rows based on a reference column and apply formulas
        to the 'Quantity' sheet starting from specified row.
        
        Args:
            reference_column: Column to check for data (default: 'D')
            start_row: Starting row number for writing formulas (default: 7)
        
        Returns:
            Dictionary with statistics about the operation
        """
        # Prefer data row count from collected CSV (constant_fill > pavement_input > tcs_input)
        from pathlib import Path as _P
        import pandas as _pd
        import os as _os
        session_data_dir = _os.getenv('SESSION_DATA_DIR')
        if session_data_dir:
            collect_base = _P(session_data_dir) / 'collected'
        else:
            collect_base = _P(_os.getcwd()) / 'data' / 'sessions' / _os.getenv('SESSION_ID', 'default') / 'collected'
        csv_candidates = [
            collect_base / 'constant_fill.csv',
            collect_base / 'pavement_input.csv',
            collect_base / 'tcs_input.csv',
            collect_base / 'tcs_schedule.csv',
        ]
        row_count = 0
        for csv_path in csv_candidates:
            if csv_path.exists():
                try:
                    row_count = len(_pd.read_csv(csv_path))
                    print(f"[FormulaApplier] Found collected CSV: {csv_path.name} with {row_count} rows")
                    break
                except Exception as e:
                    print(f"[FormulaApplier] Error reading {csv_path.name}: {e}")
                    continue
        if row_count == 0:
            print(f"[FormulaApplier] No collected CSV found in {collect_base}; listing directory...")
            if collect_base.exists():
                files = list(collect_base.glob('*.csv'))
                print(f"[FormulaApplier] Available CSVs: {[f.name for f in files]}")
            else:
                print(f"[FormulaApplier] Collected directory does not exist: {collect_base}")
            # Fallback: scan workbook
            print(f"Looking for input Excel file at: {self.input_excel_path}")
            print(f"Input Excel file exists: {self.input_excel_path.exists()}")
            if not self.input_excel_path.exists():
                alternative_path = Path(__file__).parent.parent / "data" / self.output_filename
                if alternative_path.exists():
                    self.input_excel_path = alternative_path
                else:
                    raise FileNotFoundError(f"Input Excel file not found at: {self.input_excel_path}\nAlso tried: {alternative_path}")
            input_wb = load_workbook(self.input_excel_path)
            input_sheet_name = "Quantity"
            if input_sheet_name not in input_wb.sheetnames:
                raise ValueError(f"Input sheet '{input_sheet_name}' not found in {self.input_excel_path}")
            input_sheet = input_wb[input_sheet_name]
            # Count non-empty rows from start_row in reference column
            for row_num in range(start_row, input_sheet.max_row + 1):
                val = input_sheet[f'{reference_column}{row_num}'].value
                if val is not None and str(val).strip():
                    row_count += 1
            input_wb.close()
        else:
            print(f"Found {row_count} data rows from collected CSV")
        formulas = self.template.get("formulas", {})
        if row_count == 0:
            raise ValueError("No data rows detected for formula application")
        
        # Build formula cells starting from start_row
        total_count = 0
        output_row = start_row
        cell_formulas = {}
        for i in range(row_count):
            input_row_num = start_row + i
            for col_letter, formula_template in formulas.items():
                if formula_template:
                    formula = formula_template.replace("{row}", str(input_row_num))
                    cell_formulas[f"{col_letter}{output_row}"] = formula
                    total_count += 1
            output_row += 1
        

        # Store formulas via collector for single-write later
        collector = get_collector()
        collector.add_formula_data('formula_applier', {
            'Quantity': cell_formulas
        })
        
        # Note: File will be uploaded to GCS at the end of all processing in main.py
        # No need to upload here for efficiency
        
        return {
            "input_first_row": start_row,
            "input_last_row": start_row + row_count - 1,
            "input_total_rows": row_count,
            "output_start_row": start_row,
            "output_end_row": output_row - 1,
            "formulas_per_row": len(formulas),
            "total_formulas": total_count,
            "output_file": str(self.output_excel_path),
            "output_sheet": "Quantity"
        }
    
    def apply_formulas_with_custom_mapping(self, row_mapping, start_row=7):
        """
        Apply formulas with custom row mapping from input to output.
        
        Args:
            row_mapping: Dictionary mapping input_row -> output_row
            start_row: Starting row for output (default: 7)
        
        Returns:
            Dictionary with statistics about the operation
        """
        print(f"Loading input from: {self.input_excel_path}")
        print(f"Writing output to: {self.output_excel_path}")
        
        if not self.input_excel_path.exists():
            raise FileNotFoundError(f"Input Excel file not found: {self.input_excel_path}")
        
        # Load input workbook (read-only)
        input_wb = load_workbook(self.input_excel_path)
        input_sheet_name = "Quantity"
        
        if input_sheet_name not in input_wb.sheetnames:
            raise ValueError(f"Input sheet '{input_sheet_name}' not found")
        
        input_sheet = input_wb[input_sheet_name]
        formulas = self.template.get("formulas", {})
        
        # Build formula cells using custom mapping
        total_count = 0
        cell_formulas = {}
        for input_row, output_row in row_mapping.items():
            for col_letter, formula_template in formulas.items():
                if formula_template:
                    formula = formula_template.replace("{row}", str(input_row))
                    cell_formulas[f"{col_letter}{output_row}"] = formula
                    total_count += 1
        input_wb.close()

        # Store formulas via collector
        collector = get_collector()
        collector.add_formula_data('formula_applier', {
            'Quantity': cell_formulas
        })
        
        # Note: File will be uploaded to GCS at the end of all processing in main.py
        # No need to upload here for efficiency
        
        return {
            "total_mappings": len(row_mapping),
            "total_formulas": total_count,
            "output_file": str(self.output_excel_path),
            "output_sheet": "Quantity"
        }
    
    def get_template_info(self):
        """Get information about the loaded template."""
        formulas = self.template.get("formulas", {})
        return {
            "template_name": self.template.get("template_name"),
            "source_file": self.template.get("source_file"),
            "sheet_name": self.template.get("sheet_name"),
            "source_row": self.template.get("source_row"),
            "column_range": self.template.get("column_range"),
            "total_formulas": len(formulas)
        }


def main():
    """Command line interface for applying formulas."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Apply formula templates to output/main_carriageway_and_boq.xlsx")
    parser.add_argument("--ref-column", default="D", help="Reference column for auto detection")
    parser.add_argument("--start-row", type=int, default=7, help="Starting row for output (default: 7)")
    parser.add_argument("--template", help="Path to formula template JSON")
    parser.add_argument("--input", help="Path to input main_carriageway_and_boq.xlsx file")
    parser.add_argument("--output", help="Path to output main_carriageway_and_boq.xlsx file")
    
    args = parser.parse_args()
    
    applier = FormulaApplier(template_path=args.template, 
                           input_excel_path=args.input, 
                           output_excel_path=args.output)
    
    info = applier.get_template_info()
    print(f"Loaded template: {info['template_name']}")
    print(f"Total formulas: {info['total_formulas']}")
    print(f"Column range: {info['column_range']}\n")
    
    print(f"Auto-detecting data rows from column {args.ref_column}...")
    result = applier.apply_formulas_to_all_data_rows(args.ref_column, args.start_row)
    
    print(f"✓ Input data rows: {result['input_first_row']} to {result['input_last_row']}")
    print(f"✓ Output rows: {result['output_start_row']} to {result['output_end_row']}")
    print(f"✓ Applied {result['total_formulas']} formulas to {result['input_total_rows']} rows")
    print(f"\nFormulas written to: {result['output_file']} (Sheet: {result['output_sheet']})")
    
    return 0


if __name__ == "__main__":
    exit(main())
