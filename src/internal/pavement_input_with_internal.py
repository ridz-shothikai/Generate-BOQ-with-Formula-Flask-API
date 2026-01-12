"""
Geogrid Calculator
Reads Pavement_Input.xlsx to check for Geogrid conditions
Calculates columns KY, KZ, LA, LB in main_carriageway.xlsx based on formulas
"""

import pandas as pd
import os
import sys
import io
from dotenv import load_dotenv

# Add project root to Python path
project_root = os.path.join(os.path.dirname(__file__), '..', '..')
sys.path.append(project_root)

from src.utils.gcs_utils import get_gcs_handler

load_dotenv()
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


# ============================================================================
# FILE PATHS - Update these to match your folder structure
# ============================================================================

# NEW CODE:
script_dir = os.path.dirname(os.path.abspath(__file__))
session_id = os.getenv('SESSION_ID', 'default')
is_merged = os.getenv('IS_MERGED', 'True').lower() == 'true'

# Use SESSION_OUTPUT_FILE if available (local file), otherwise fallback to GCS
output_file = os.getenv('SESSION_OUTPUT_FILE', '')
if output_file and os.path.exists(output_file):
    MAIN_CARRIAGEWAY_FILE = output_file
    OUTPUT_EXCEL = output_file
    print(f"Using local output file: {output_file}")
else:
    # Fallback: download from GCS (for backward compatibility)
    gcs = get_gcs_handler()
    if is_merged:
        output_filename = f"{session_id}_main_carriageway_and_boq.xlsx"
    else:
        output_filename = f"{session_id}_main_carriageway.xlsx"
    output_gcs_path = gcs.get_gcs_path(session_id, output_filename, 'output')
    temp_main_file = gcs.download_to_temp(output_gcs_path, suffix='.xlsx')
    MAIN_CARRIAGEWAY_FILE = temp_main_file
    OUTPUT_EXCEL = temp_main_file
    print(f"[GCS] Downloaded output file from GCS: {temp_main_file}")

# Initialize GCS for input files only
gcs = get_gcs_handler()

# Download input files from GCS
pavement_gcs = gcs.get_gcs_path(session_id, 'Pavement Input.xlsx', 'data')
PAVEMENT_INPUT_FILE = gcs.download_to_temp(pavement_gcs, suffix='.xlsx')


# ============================================================================
# STEP 1: Check Pavement_Input.xlsx for Geogrid Conditions
# ============================================================================

def check_geogrid_conditions(pavement_input_file):
    """
    Reads Pavement_Input.xlsx and checks if:
    - E9 contains "Geogrid Reinforced GSB"
    - E10 contains "Geogrid Reinforced WMM"
    - B9 contains "Geogrid Reinforced GSB"
    - B10 contains "Geogrid Reinforced WMM"
    Returns: dictionary with boolean flags
    """
    print("="*80)
    print("STEP 1: Checking Geogrid Conditions in Pavement Input")
    print("="*80)
    
    # Read the Excel file without headers
    df = pd.read_excel(pavement_input_file, header=None)
    
    print("[OK] Read Pavement_Input.xlsx:", len(df), "total rows")
    
    # Initialize conditions
    conditions = {
        'e9_geogrid_gsb': False,
        'e10_geogrid_wmm': False,
        'b9_geogrid_gsb': False,
        'b10_geogrid_wmm': False
    }
    
    # Check E9 (row index 8, column 4)
    if len(df) > 8:
        e9_value = df.iloc[8, 4]  # E9
        if pd.notna(e9_value) and "Geogrid Reinforced GSB" in str(e9_value):
            conditions['e9_geogrid_gsb'] = True
            print(f"[OK] E9 contains 'Geogrid Reinforced GSB': {e9_value}")
        else:
            print(f"[OK] E9 value: {e9_value} (not Geogrid Reinforced GSB)")
    
    # Check E10 (row index 9, column 4)
    if len(df) > 9:
        e10_value = df.iloc[9, 4]  # E10
        if pd.notna(e10_value) and "Geogrid Reinforced WMM" in str(e10_value):
            conditions['e10_geogrid_wmm'] = True
            print(f"[OK] E10 contains 'Geogrid Reinforced WMM': {e10_value}")
        else:
            print(f"[OK] E10 value: {e10_value} (not Geogrid Reinforced WMM)")
    
    # Check B9 (row index 8, column 1)
    if len(df) > 8:
        b9_value = df.iloc[8, 1]  # B9
        if pd.notna(b9_value) and "Geogrid Reinforced GSB" in str(b9_value):
            conditions['b9_geogrid_gsb'] = True
            print(f"[OK] B9 contains 'Geogrid Reinforced GSB': {b9_value}")
        else:
            print(f"[OK] B9 value: {b9_value} (not Geogrid Reinforced GSB)")
    
    # Check B10 (row index 9, column 1)
    if len(df) > 9:
        b10_value = df.iloc[9, 1]  # B10
        if pd.notna(b10_value) and "Geogrid Reinforced WMM" in str(b10_value):
            conditions['b10_geogrid_wmm'] = True
            print(f"[OK] B10 contains 'Geogrid Reinforced WMM': {b10_value}")
        else:
            print(f"[OK] B10 value: {b10_value} (not Geogrid Reinforced WMM)")
    
    print("\n[OK] Geogrid Conditions Summary:")
    print(f"  E9 Geogrid GSB: {conditions['e9_geogrid_gsb']}")
    print(f"  E10 Geogrid WMM: {conditions['e10_geogrid_wmm']}")
    print(f"  B9 Geogrid GSB: {conditions['b9_geogrid_gsb']}")
    print(f"  B10 Geogrid WMM: {conditions['b10_geogrid_wmm']}")
    
    return conditions


# ============================================================================
# STEP 2: Calculate Geogrid Columns
# ============================================================================

def find_last_row_with_data(ws, column_letter):
    """
    Find the last row with data in the specified column
    Args:
        ws: Worksheet object
        column_letter: Excel column letter (e.g., 'D')
    Returns:
        Last row number that contains data
    """
    from openpyxl.utils import column_index_from_string
    
    col_idx = column_index_from_string(column_letter)
    last_row = 0
    
    # Iterate from bottom to top to find last non-empty cell
    for row_idx in range(ws.max_row, 0, -1):
        cell_value = ws.cell(row_idx, col_idx).value
        if cell_value is not None and cell_value != '':
            last_row = row_idx
            break
    
    return last_row

def calculate_geogrid_columns(main_carriageway_file, conditions, output_file):
    """
    Reads main_carriageway_and_boq.xlsx and calculates geogrid columns based on conditions
    Updates only columns KY, KZ, LA, LB without touching other columns
    """
    print("\n" + "="*80)
    print("STEP 2: Calculating Geogrid Columns")
    print("="*80)
    
    from openpyxl import load_workbook
    
    # Load workbook normally (not data_only) to preserve formulas when saving
    wb = load_workbook(main_carriageway_file)
    ws = wb['Quantity']
    
    print(f"[OK] Loaded workbook: {main_carriageway_file}")
    print("  Sheet: Quantity")
    print(f"  Max row: {ws.max_row}, Max column: {ws.max_column}")
    
    # Find last row with data using column D as reference
    last_row_with_data = find_last_row_with_data(ws, 'D')
    print(f"[OK] Last row with data in column D: {last_row_with_data}")
    
    if last_row_with_data == 0:
        print("[WARNING] No data found in column D, using max_row instead")
        last_row_with_data = ws.max_row
    
    # Column letters (Excel columns, 1-indexed)
    LENGTH_COL = 3      # Column C
    DL_COL = 116        # Column DL
    DS_COL = 123        # Column DS
    EE_COL = 135        # Column EE
    EJ_COL = 140        # Column EJ
    FD_COL = 160        # Column FD
    FF_COL = 162        # Column FF
    FS_COL = 175        # Column FS
    FY_COL = 181        # Column FY
    
    KY_COL = 311        # Column KY
    KZ_COL = 312        # Column KZ
    LA_COL = 313        # Column LA
    LB_COL = 314        # Column LB
    
    # Data starts from row 7
    start_row = 7
    
    print(f"\n[OK] Calculating geogrid values from row {start_row} to row {last_row_with_data}...")
    
    # Helper to safely coerce values to float (handles None, blanks, numeric strings, formulas)
    def to_float(val):
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        # If it's a string (including formula strings like "=SUM(...)"), try to parse
        # For formula strings, this will fail and return 0.0, which is safe for our calculation
        try:
            s = str(val).strip().replace(',', '')
            if not s:
                return 0.0
            # Skip formula strings - they haven't been calculated yet
            if s.startswith('='):
                return 0.0
            return float(s)
        except (ValueError, TypeError):
            return 0.0
    
    # OPTIMIZATION: Pre-read all column data into lists (much faster than cell-by-cell access)
    print("[OPTIMIZATION] Reading column data...")
    length_vals = [to_float(ws.cell(r, LENGTH_COL).value) for r in range(start_row, last_row_with_data + 1)]
    dl_vals = [to_float(ws.cell(r, DL_COL).value) for r in range(start_row, last_row_with_data + 1)]
    ds_vals = [to_float(ws.cell(r, DS_COL).value) for r in range(start_row, last_row_with_data + 1)]
    ee_vals = [to_float(ws.cell(r, EE_COL).value) for r in range(start_row, last_row_with_data + 1)]
    ej_vals = [to_float(ws.cell(r, EJ_COL).value) for r in range(start_row, last_row_with_data + 1)]
    ff_vals = [to_float(ws.cell(r, FF_COL).value) for r in range(start_row, last_row_with_data + 1)]
    fd_vals = [to_float(ws.cell(r, FD_COL).value) for r in range(start_row, last_row_with_data + 1)]
    fy_vals = [to_float(ws.cell(r, FY_COL).value) for r in range(start_row, last_row_with_data + 1)]
    fs_vals = [to_float(ws.cell(r, FS_COL).value) for r in range(start_row, last_row_with_data + 1)]
    
    # Pre-compute condition multipliers (avoid repeated condition checks in loop)
    e9_mult = 1 if conditions['e9_geogrid_gsb'] else 0
    e10_mult = 1 if conditions['e10_geogrid_wmm'] else 0
    b9_mult = 1 if conditions['b9_geogrid_gsb'] else 0
    b10_mult = 1 if conditions['b10_geogrid_wmm'] else 0
    
    # Batch calculate all values
    print(f"[OPTIMIZATION] Calculating {len(length_vals)} rows...")
    ky_results = []
    kz_results = []
    la_results = []
    lb_results = []
    row_count = 0
    
    for i, length in enumerate(length_vals):
        # Skip empty rows
        if length == 0 or length is None or length == '':
            ky_results.append(None)
            kz_results.append(None)
            la_results.append(None)
            lb_results.append(None)
            continue
        
        # Values are already coerced to float
        dl_val = dl_vals[i]
        ds_val = ds_vals[i]
        ee_val = ee_vals[i]
        ej_val = ej_vals[i]
        ff_val = ff_vals[i]
        fd_val = fd_vals[i]
        fy_val = fy_vals[i]
        fs_val = fs_vals[i]
        
        # Calculate values using pre-computed multipliers
        ky_val = (dl_val * float(e9_mult) + ds_val * float(e10_mult)) * length
        kz_val = (ee_val * float(b9_mult) + ej_val * float(b10_mult)) * length
        la_val = (ff_val * float(b9_mult) + fd_val * float(b10_mult)) * length
        lb_val = (fy_val * float(e9_mult) + fs_val * float(e10_mult)) * length
        
        ky_results.append(ky_val)
        kz_results.append(kz_val)
        la_results.append(la_val)
        lb_results.append(lb_val)
        row_count += 1
    
    # Batch-write all results
    print(f"[OPTIMIZATION] Writing {row_count} rows to worksheet...")
    for i, row_idx in enumerate(range(start_row, last_row_with_data + 1)):
        if ky_results[i] is not None:
            ws.cell(row_idx, KY_COL).value = ky_results[i]
            ws.cell(row_idx, KZ_COL).value = kz_results[i]
            ws.cell(row_idx, LA_COL).value = la_results[i]
            ws.cell(row_idx, LB_COL).value = lb_results[i]
    
    print(f"[OK] Geogrid calculations completed for {row_count} rows")
    
    # Save workbook
    print(f"\n[OK] Saving to {output_file}...")
    wb.save(output_file)
    
    print("[OK] Saved!")
    
    return row_count


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main function to execute all steps"""
    print("\n" + "="*80)
    print("GEOGRID CALCULATOR")
    print("="*80)
    print("Configuration:")
    print("  • Pavement_Input.xlsx: Check for Geogrid conditions")
    print("  • main_carriageway_and_boq.xlsx: Calculate columns KY, KZ, LA, LB")
    print("="*80 + "\n")
    
    try:
        # Step 1: Check geogrid conditions
        conditions = check_geogrid_conditions(PAVEMENT_INPUT_FILE)
        
        # Step 2: Calculate geogrid columns
        row_count = calculate_geogrid_columns(MAIN_CARRIAGEWAY_FILE, conditions, OUTPUT_EXCEL)
        
        # Success summary
        print("\n" + "="*80)
        print("SUCCESS! [OK]")
        print("="*80)
        print("Output file:", OUTPUT_EXCEL)
        print("Total rows processed:", row_count)
        print("Columns updated: KY, KZ, LA, LB")
        
        # Note: File will be uploaded to GCS at the end of all processing in main.py
        # No need to upload here for efficiency
        
        # Cleanup - only remove temp files, not the local output file
        os.remove(PAVEMENT_INPUT_FILE)
        # Only remove OUTPUT_EXCEL if it's a temp file (not the local SESSION_OUTPUT_FILE)
        if OUTPUT_EXCEL != os.getenv('SESSION_OUTPUT_FILE', ''):
            try:
                os.remove(OUTPUT_EXCEL)
            except:
                pass
        
    except FileNotFoundError as e:
        print("\n[ERROR] File not found")
        print(" ", e)
        print("\nPlease check:")
        print("  1. Files exist in the data folder")
        print("  2. File names match exactly")
    except Exception as e:
        print("\n[ERROR]:", e)
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
