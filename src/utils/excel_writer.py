"""
Excel Writer Module
Writes all collected data to Excel in a single operation
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import sys
import os
import shutil

# Add project root to Python path
project_root = os.path.join(os.path.dirname(__file__), '..', '..')
sys.path.append(project_root)

from src.utils.data_collector import get_collector
import json


class ExcelWriter:
    """Handles writing all collected data to Excel in one operation"""
    
    def __init__(self, output_file_path):
        """
        Initialize Excel writer
        
        Args:
            output_file_path (str): Path to the output Excel file
        """
        self.output_file = Path(output_file_path)
        self.collector = get_collector()
        
        if not self.output_file.exists():
            raise FileNotFoundError(f"Output file not found: {self.output_file}")
    
    def write_all_data(self):
        """Write all collected data to Excel file in one operation"""
        print("\n" + "="*80)
        print("WRITING ALL DATA TO EXCEL")
        print("="*80)
        print(f"Output file: {self.output_file}")
        
        # Load manifest from collector (disk-backed)
        all_data = self.collector.get_all_data()
        
        # Load workbook once
        wb = openpyxl.load_workbook(self.output_file)
        
        # Process each script's data in sequence
        scripts_to_process = list(all_data.get('scripts', {}).keys()) or [
            'tcs_schedule', 'tcs_input', 'emb_height', 'pavement_input',
            'constant_fill', 'formula_applier', 'pavement_input_with_internal', 'final_sum_applier']
        
        for script_name in scripts_to_process:
            script_entry = all_data.get('scripts', {}).get(script_name)
            
            if script_entry is None:
                print(f"[ExcelWriter] No data for {script_name}, skipping")
                continue
            
            print(f"\n[ExcelWriter] Processing {script_name}...")
            
            try:
                # Handle different data types
                data_type = script_entry.get('type', 'dataframe')
                
                if data_type == 'dataframe':
                    self._write_dataframe_data_from_manifest(wb, script_name, script_entry)
                elif data_type == 'formulas':
                    self._write_formula_data_from_manifest(wb, script_name, script_entry)
                elif data_type == 'modifications':
                    self._write_modifications_from_manifest(wb, script_name, script_entry)
                else:
                    print(f"[ExcelWriter] Unknown data type for {script_name}: {data_type}")
                
                print(f"[ExcelWriter] ✓ Completed {script_name}")
                
            except Exception as e:
                print(f"[ExcelWriter] ✗ Error processing {script_name}: {str(e)}")
                raise
        
        # Save workbook once at the end
        print(f"\n[ExcelWriter] Saving workbook to {self.output_file}...")
        wb.save(self.output_file)
        print(f"[ExcelWriter] ✓ Successfully saved workbook")
        
        print("="*80)
        print("DATA WRITING COMPLETED")
        print("="*80)
    
    def cleanup_collected_data(self):
        """Delete the session's collected temp files and directory.

        Honors env var KEEP_COLLECTED (true/1/yes to skip deletion).
        Also removes the session directory if empty, or forcibly when FORCE_SESSION_CLEAN=true.
        """
        keep = os.getenv('KEEP_COLLECTED', 'false').lower() in ('1', 'true', 'yes')
        if keep:
            print("[ExcelWriter] KEEP_COLLECTED is true; skipping cleanup of collected temp files")
            return

        session_data_dir = os.getenv('SESSION_DATA_DIR')
        if session_data_dir:
            session_dir = Path(session_data_dir)
            collect_dir = session_dir / 'collected'
        else:
            session_dir = Path(os.getcwd()) / 'data' / 'sessions' / os.getenv('SESSION_ID', 'default')
            collect_dir = session_dir / 'collected'

        if not collect_dir.exists():
            print(f"[ExcelWriter] Collected directory not found (nothing to clean): {collect_dir}")
        else:
            try:
                print(f"[ExcelWriter] Cleaning collected temp files at {collect_dir}...")
                shutil.rmtree(collect_dir)
                print("[ExcelWriter] ✓ Collected temp files removed")
            except Exception as e:
                print(f"[ExcelWriter] ✗ Failed to remove collected temp files: {e}")

        # Attempt to remove the session directory as requested
        keep_session_dir = os.getenv('KEEP_SESSION_DIR', 'false').lower() in ('1', 'true', 'yes')
        force_session_clean = os.getenv('FORCE_SESSION_CLEAN', 'false').lower() in ('1', 'true', 'yes')
        if keep_session_dir:
            print(f"[ExcelWriter] KEEP_SESSION_DIR is true; skipping session folder cleanup: {session_dir}")
            return

        if not session_dir.exists():
            print(f"[ExcelWriter] Session folder not found (nothing to clean): {session_dir}")
            return

        try:
            if force_session_clean:
                print(f"[ExcelWriter] FORCE_SESSION_CLEAN is true; removing session folder: {session_dir}")
                shutil.rmtree(session_dir)
                print("[ExcelWriter] ✓ Session folder removed")
            else:
                # Remove only if empty
                if not any(session_dir.iterdir()):
                    session_dir.rmdir()
                    print("[ExcelWriter] ✓ Session folder removed (was empty)")
                else:
                    print(f"[ExcelWriter] Session folder not empty; skipping removal: {session_dir}")
        except Exception as e:
            print(f"[ExcelWriter] ✗ Failed to remove session folder: {e}")

    def _write_dataframe_data(self, wb, script_name, script_data):
        """Write DataFrame data using pandas ExcelWriter"""
        sheet_name = script_data.get('sheet_name')
        df = script_data.get('dataframe')
        start_row = script_data.get('start_row', 0)
        start_col = script_data.get('start_col', 0)
        write_header = script_data.get('write_header', False)
        write_index = script_data.get('write_index', False)
        
        if df is None or sheet_name is None:
            print(f"[ExcelWriter] Missing dataframe or sheet_name for {script_name}")
            return
        
        # Save current workbook state
        wb.save(self.output_file)
        
        # Use pandas ExcelWriter to write DataFrame
        with pd.ExcelWriter(self.output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=start_row,
                startcol=start_col,
                index=write_index,
                header=write_header
            )
        
        # Reload workbook after pandas write
        wb_new = openpyxl.load_workbook(self.output_file)
        
        # Update the reference (copy sheets)
        for sheet_name_iter in wb_new.sheetnames:
            if sheet_name_iter in wb.sheetnames:
                del wb[sheet_name_iter]
            wb._sheets.append(wb_new[sheet_name_iter])
        
        print(f"[ExcelWriter]   Written DataFrame to {sheet_name} at row {start_row+1}, col {start_col+1}")
        print(f"[ExcelWriter]   DataFrame shape: {df.shape}")

    def _write_dataframe_data_from_manifest(self, wb, script_name, script_entry):
        """Load DataFrame from CSV per manifest and write"""
        import os as _os
        session_data_dir = _os.getenv('SESSION_DATA_DIR')
        if session_data_dir:
            collect_dir = Path(session_data_dir) / 'collected'
        else:
            collect_dir = Path(_os.getcwd()) / 'data' / 'sessions' / _os.getenv('SESSION_ID', 'default') / 'collected'
        data_file = collect_dir / script_entry.get('data_file')
        if not data_file.exists():
            print(f"[ExcelWriter] Data file not found for {script_name}: {data_file}")
            return
        df = pd.read_csv(data_file)
        self._write_dataframe_data(wb, script_name, {
            'sheet_name': script_entry.get('sheet_name'),
            'dataframe': df,
            'start_row': script_entry.get('start_row', 0),
            'start_col': script_entry.get('start_col', 0),
            'write_header': script_entry.get('write_header', False),
            'write_index': script_entry.get('write_index', False)
        })
    
    def _write_formula_data(self, wb, script_name, script_data):
        """Write formula data to Excel"""
        formula_data = script_data.get('data', {})
        
        for sheet_name, cell_formulas in formula_data.items():
            if sheet_name not in wb.sheetnames:
                print(f"[ExcelWriter] Sheet {sheet_name} not found, skipping")
                continue
            
            ws = wb[sheet_name]
            formula_count = 0
            
            for cell_ref, formula in cell_formulas.items():
                ws[cell_ref] = formula
                formula_count += 1
            
            print(f"[ExcelWriter]   Applied {formula_count} formulas to sheet {sheet_name}")

    def _write_formula_data_from_manifest(self, wb, script_name, script_entry):
        import os as _os
        session_data_dir = _os.getenv('SESSION_DATA_DIR')
        if session_data_dir:
            collect_dir = Path(session_data_dir) / 'collected'
        else:
            collect_dir = Path(_os.getcwd()) / 'data' / 'sessions' / _os.getenv('SESSION_ID', 'default') / 'collected'
        data_file = collect_dir / script_entry.get('data_file')
        if not data_file.exists():
            print(f"[ExcelWriter] Formula file not found for {script_name}: {data_file}")
            return
        with open(data_file, 'r') as f:
            formula_data = json.load(f)
        self._write_formula_data(wb, script_name, {'data': formula_data})
    
    def _write_modifications(self, wb, script_name, script_data):
        """Write cell modifications to Excel"""
        modifications = script_data.get('data', [])
        
        for mod in modifications:
            sheet_name = mod.get('sheet')
            
            if sheet_name not in wb.sheetnames:
                print(f"[ExcelWriter] Sheet {sheet_name} not found, skipping")
                continue
            
            ws = wb[sheet_name]
            
            # Support both cell reference and row/col coordinates
            if 'cell' in mod:
                cell_ref = mod['cell']
                cell = ws[cell_ref]
            elif 'row' in mod and 'col' in mod:
                row = mod['row']
                col = mod['col']
                cell = ws.cell(row=row, column=col)
            else:
                print(f"[ExcelWriter] Invalid modification: {mod}")
                continue
            
            # Write value or formula
            if 'value' in mod:
                cell.value = mod['value']
            elif 'formula' in mod:
                cell.value = mod['formula']
            
            # Apply formatting if specified
            if 'number_format' in mod:
                cell.number_format = mod['number_format']

    def _write_modifications_from_manifest(self, wb, script_name, script_entry):
        import os as _os
        session_data_dir = _os.getenv('SESSION_DATA_DIR')
        if session_data_dir:
            collect_dir = Path(session_data_dir) / 'collected'
        else:
            collect_dir = Path(_os.getcwd()) / 'data' / 'sessions' / _os.getenv('SESSION_ID', 'default') / 'collected'
        data_file = collect_dir / script_entry.get('data_file')
        if not data_file.exists():
            print(f"[ExcelWriter] Modifications file not found for {script_name}: {data_file}")
            return
        with open(data_file, 'r') as f:
            modifications = json.load(f)
        self._write_modifications(wb, script_name, {'data': modifications})


def write_all_collected_data(output_file_path):
    """
    Convenience function to write all collected data
    
    Args:
        output_file_path (str): Path to the output Excel file
    """
    writer = ExcelWriter(output_file_path)
    writer.write_all_data()
    # Attempt cleanup by default (toggle with KEEP_COLLECTED)
    writer.cleanup_collected_data()


if __name__ == "__main__":
    # For testing
    output_file = os.getenv('SESSION_OUTPUT_FILE')
    if output_file:
        write_all_collected_data(output_file)
    else:
        print("Error: SESSION_OUTPUT_FILE environment variable not set")
        sys.exit(1)
